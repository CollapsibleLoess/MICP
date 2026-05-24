import dearpygui.dearpygui as dpg
import numpy as np
import sys, os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from core import MICPProcessor, export_excel, export_json

C_CYAN = [0, 200, 220]
C_RED = [233, 69, 96]
C_GOLD = [254, 202, 87]
C_PURPLE = [140, 80, 220]
C_DIM = [140, 140, 160]

CHARTS = [
    ("capillary",      "\u6bdb\u7ba1\u538b\u529b\u66f2\u7ebf"),
    ("dvdD",           "\u5b54\u5f84\u5206\u5e03(dV/dD)"),
    ("dvdlogD",        "\u5b54\u5f84\u5206\u5e03(dV/dlogD)"),
    ("pct",            "\u5b54\u5f84\u5206\u5e03(%)"),
    ("inj_ext",        "\u8fdb\u9000\u6c55\u66f2\u7ebf"),
    ("characteristic", "\u7279\u5f81\u53c2\u6570\u5206\u5e03"),
    ("fractal",        "\u5206\u5f62\u7ef4\u6570"),
    ("mf_spectrum",    "\u591a\u91cd\u5206\u5f62\u8c31f(\u03b1)"),
    ("mf_Dq",          "\u5e7f\u4e49\u7ef4\u6570D(q)"),
    ("mf_tau",         "\u8d28\u91cf\u6807\u5ea6\u6307\u6570\u03c4(q)"),
    ("ratio",          "\u5b54\u5589\u6bd4"),
    ("swanson",        "Swanson\u6e17\u900f\u7387"),
]


class App:
    def __init__(self):
        self.processor = None
        self.result = None
        self.batch_results = []
        self.yax = None
        self.current_chart = CHARTS[0][0]
        self.res_tags = {}

    def run(self):
        dpg.create_context()
        self._setup_theme()
        self._build_ui()
        dpg.create_viewport(
            title="MICP \u538b\u6c55\u6cd5\u6570\u636e\u5904\u7406\u7cfb\u7edf",
            width=1440, height=900,
            clear_color=[18, 18, 30]
        )
        dpg.setup_dearpygui()
        dpg.show_viewport()
        dpg.set_primary_window("win", True)
        dpg.start_dearpygui()
        dpg.destroy_context()

    def _setup_theme(self):
        with dpg.theme() as t:
            with dpg.theme_component(dpg.mvAll):
                dpg.add_theme_color(dpg.mvThemeCol_WindowBg,       [18, 18, 30, 255])
                dpg.add_theme_color(dpg.mvThemeCol_ChildBg,        [22, 28, 48, 255])
                dpg.add_theme_color(dpg.mvThemeCol_PopupBg,        [22, 28, 48, 255])
                dpg.add_theme_color(dpg.mvThemeCol_Border,         [40, 50, 80, 180])
                dpg.add_theme_color(dpg.mvThemeCol_FrameBg,        [28, 28, 46, 255])
                dpg.add_theme_color(dpg.mvThemeCol_FrameBgHovered, [36, 36, 60, 255])
                dpg.add_theme_color(dpg.mvThemeCol_FrameBgActive,  [44, 44, 74, 255])
                dpg.add_theme_color(dpg.mvThemeCol_Button,         [15, 52, 96, 255])
                dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered,  [83, 52, 131, 255])
                dpg.add_theme_color(dpg.mvThemeCol_ButtonActive,   [0, 160, 180, 255])
                dpg.add_theme_color(dpg.mvThemeCol_Text,           [230, 230, 240, 255])
                dpg.add_theme_color(dpg.mvThemeCol_TextDisabled,   [120, 120, 140, 255])
                dpg.add_theme_color(dpg.mvThemeCol_Tab,            [22, 28, 48, 255])
                dpg.add_theme_color(dpg.mvThemeCol_TabHovered,     [15, 52, 96, 255])
                dpg.add_theme_color(dpg.mvThemeCol_TabActive,      [0, 160, 180, 255])
                dpg.add_theme_color(dpg.mvThemeCol_Separator,      [40, 50, 80, 180])
                dpg.add_theme_color(dpg.mvThemeCol_TitleBg,        [15, 15, 28, 255])
                dpg.add_theme_color(dpg.mvThemeCol_TitleBgActive,  [15, 52, 96, 255])
                dpg.add_theme_color(dpg.mvThemeCol_MenuBarBg,      [15, 15, 28, 255])
                dpg.add_theme_color(dpg.mvThemeCol_CheckMark,      [0, 200, 220, 255])
                dpg.add_theme_color(dpg.mvThemeCol_Header,         [15, 52, 96, 255])
                dpg.add_theme_color(dpg.mvThemeCol_HeaderHovered,  [83, 52, 131, 255])
                dpg.add_theme_color(dpg.mvThemeCol_HeaderActive,   [0, 160, 180, 255])
                dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 5)
                dpg.add_theme_style(dpg.mvStyleVar_GrabRounding, 4)
                dpg.add_theme_style(dpg.mvStyleVar_WindowRounding, 6)
                dpg.add_theme_style(dpg.mvStyleVar_ChildRounding, 6)
                dpg.add_theme_style(dpg.mvStyleVar_TabRounding, 4)
                dpg.add_theme_style(dpg.mvStyleVar_FramePadding, 6, 3)
        dpg.bind_theme(t)

        with dpg.theme(tag="_pt"):
            with dpg.theme_component(dpg.mvPlot):
                dpg.add_theme_color(dpg.mvPlotCol_PlotBg,       [12, 12, 22, 255])
                dpg.add_theme_color(dpg.mvPlotCol_PlotBorder,   [40, 50, 80, 180])
                dpg.add_theme_color(dpg.mvPlotCol_XAxis,        [140, 140, 160, 255])
                dpg.add_theme_color(dpg.mvPlotCol_YAxis,        [140, 140, 160, 255])
                dpg.add_theme_color(dpg.mvPlotCol_XAxisGrid,    [40, 50, 80, 100])
                dpg.add_theme_color(dpg.mvPlotCol_YAxisGrid,    [40, 50, 80, 100])
                dpg.add_theme_color(dpg.mvPlotCol_LegendBg,     [22, 28, 48, 220])
                dpg.add_theme_color(dpg.mvPlotCol_LegendBorder, [40, 50, 80, 180])
                dpg.add_theme_color(dpg.mvPlotCol_LegendText,   [210, 210, 220, 255])

    def _build_ui(self):
        with dpg.window(tag="win"):
            with dpg.menu_bar():
                with dpg.menu(label="\u6587\u4ef6"):
                    dpg.add_menu_item(label="\u6253\u5f00\u6587\u4ef6", callback=self._show_open)
                    dpg.add_menu_item(label="\u6279\u91cf\u5904\u7406", callback=self._show_batch)
                    dpg.add_separator()
                    dpg.add_menu_item(label="\u9000\u51fa", callback=lambda: dpg.stop_dearpygui())
                with dpg.menu(label="\u5bfc\u51fa"):
                    dpg.add_menu_item(label="Excel \u62a5\u544a", callback=self._do_export_xlsx)
                    dpg.add_menu_item(label="JSON \u6570\u636e", callback=self._do_export_json)

            with dpg.group(horizontal=True):
                with dpg.child_window(width=360, tag="left"):
                    self._build_left()
                with dpg.child_window(tag="right"):
                    with dpg.group(horizontal=True, parent="right"):
                        dpg.add_text("\u56fe\u8868:", parent="right")
                        dpg.add_combo(
                            [c[1] for c in CHARTS],
                            default_value=CHARTS[0][1],
                            tag="chart_sel", width=220,
                            callback=self._on_chart_sel,
                            parent="right"
                        )
                    with dpg.group(tag="plot_area", parent="right"):
                        self._create_plot()
                    dpg.add_separator(parent="right")
                    self._build_results(parent="right")

        with dpg.file_dialog(
            label="\u6253\u5f00\u538b\u6c55\u6570\u636e\u6587\u4ef6",
            directory_selector=False, show=False,
            callback=self._on_file, tag="fd_open", width=800, height=500
        ):
            dpg.add_file_extension(".xlsx", label="Excel")
            dpg.add_file_extension(".xls", label="Excel 97")
            dpg.add_file_extension(".xlsm", label="Macro Excel")
            dpg.add_file_extension(".*", label="\u6240\u6709\u6587\u4ef6")

        with dpg.file_dialog(
            label="\u9009\u62e9\u6587\u4ef6\u5939",
            directory_selector=True, show=False,
            callback=self._on_batch, tag="fd_batch", width=800, height=500
        ):
            pass

        with dpg.file_dialog(
            label="\u4fdd\u5b58 Excel",
            directory_selector=False, show=False,
            callback=self._on_save_xlsx, tag="fd_sxlsx", width=800, height=500
        ):
            dpg.add_file_extension(".xlsx", label="Excel")

        with dpg.file_dialog(
            label="\u4fdd\u5b58 JSON",
            directory_selector=False, show=False,
            callback=self._on_save_json, tag="fd_sjson", width=800, height=500
        ):
            dpg.add_file_extension(".json", label="JSON")

    def _build_left(self):
        dpg.add_spacer(height=4, parent="left")
        dpg.add_text("MICP Processor", color=C_RED, parent="left")
        dpg.add_text("\u538b\u6c55\u6cd5\u6bdb\u7ba1\u538b\u529b\u66f2\u7ebf\u6570\u636e\u5904\u7406", color=C_DIM, parent="left")
        dpg.add_spacer(height=6, parent="left")
        dpg.add_separator(parent="left")

        dpg.add_text("\u6587\u4ef6\u64cd\u4f5c", color=C_CYAN, parent="left")
        with dpg.group(horizontal=True, parent="left"):
            dpg.add_button(label="\u6253\u5f00\u6587\u4ef6", callback=self._show_open, width=162, height=30)
            dpg.add_button(label="\u6279\u91cf\u5904\u7406", callback=self._show_batch, width=162, height=30)
        dpg.add_text("\u672a\u52a0\u8f7d\u6587\u4ef6", tag="file_lbl", color=C_DIM, parent="left")
        dpg.add_spacer(height=4, parent="left")
        dpg.add_separator(parent="left")

        dpg.add_text("\u5b9e\u9a8c\u53c2\u6570", color=C_CYAN, parent="left")
        for key, lbl, val, dtype in [
            ("contact_angle", "\u63a5\u89e6\u89d2 (\u00b0)", 130.0, "f"),
            ("surface_tension", "\u8868\u9762\u5f20\u529b (N/m)", 0.475, "f"),
            ("mercury_pressure", "\u8fdb\u6c55\u8d77\u59cb\u538b\u529b (MPa)", 0.6, "f"),
            ("withdrawal_coeff", "\u9000\u6c55\u7cfb\u6570", 0.2, "f"),
            ("smoothing", "\u5e73\u6ed1\u7a97\u53e3", 7, "i"),
        ]:
            with dpg.group(horizontal=True, parent="left"):
                dpg.add_text(lbl, color=[190, 190, 205], parent="left")
                if dtype == "f":
                    dpg.add_input_float(tag=f"p_{key}", default_value=val, width=110, format="%.4f", step=0, parent="left")
                else:
                    dpg.add_input_int(tag=f"p_{key}", default_value=val, width=110, parent="left")
        dpg.add_spacer(height=2, parent="left")
        dpg.add_button(label="\u91cd\u65b0\u8ba1\u7b97", callback=self._recalc, width=-1, height=30, parent="left")
        dpg.add_separator(parent="left")

        dpg.add_text("\u6837\u54c1\u4fe1\u606f", color=C_CYAN, parent="left")
        dpg.add_text("\u52a0\u8f7d\u6570\u636e\u540e\u663e\u793a...", tag="sample_txt", color=C_DIM, parent="left")
        dpg.add_separator(parent="left")

        dpg.add_text("\u5bfc\u51fa", color=C_CYAN, parent="left")
        with dpg.group(horizontal=True, parent="left"):
            dpg.add_button(label="\u5bfc\u51fa Excel", callback=self._do_export_xlsx, width=162, height=30)
            dpg.add_button(label="\u5bfc\u51fa JSON", callback=self._do_export_json, width=162, height=30)

    def _build_results(self, parent):
        with dpg.collapsing_header(label="\u8ba1\u7b97\u7ed3\u679c", default_open=True, parent=parent, tag="res_header"):
            sections = [
                ("\u57fa\u672c\u53c2\u6570", [
                    ("he_por", "\u6c25\u5b54\u9699\u5ea6(%)"), ("micp_por", "\u538b\u6c55\u5b54\u9699\u5ea6(%)"),
                    ("bulk_d", "\u4f53\u79ef\u5bc6\u5ea6(g/cm\u00b3)"), ("skel_d", "\u9aa8\u67b6\u5bc6\u5ea6(g/cm\u00b3)"),
                    ("ssa", "\u6bd4\u8868\u9762\u79ef(m\u00b2/g)"),
                ]),
                ("\u7ed3\u6784\u53c2\u6570", [
                    ("inj_sat", "\u8fdb\u6c55\u9971\u548c\u5ea6(%)"), ("eff", "\u9000\u6c55\u6548\u7387(%)"),
                    ("disp_p", "\u6392\u9a71\u538b\u529b(MPa)"), ("max_d", "\u6700\u5927\u5b54\u5f84(\u03bcm)"),
                    ("med_p", "\u4e2d\u503c\u538b\u529b(MPa)"), ("med_d", "\u4e2d\u503c\u5b54\u5f84(\u03bcm)"),
                    ("pore_v", "\u5b54\u9699\u4f53\u79ef(mL/g)"),
                ]),
                ("\u7279\u5f81\u53c2\u6570", [
                    ("sp", "\u5206\u9009\u7cfb\u6570Sp"), ("skp", "\u6b6a\u5ea6Skp"),
                    ("kp", "\u5cf0\u6001Kp"), ("dm", "\u534a\u5f84\u5747\u503cDM"),
                    ("phi", "\u7ed3\u6784\u7cfb\u6570\u03d5"), ("d_coeff", "\u76f8\u5bf9\u5206\u9009D"),
                    ("frac_d", "\u5206\u5f62\u7ef4\u6570"),
                ]),
                ("\u6e17\u900f\u7387", [
                    ("k413", "\u57fa\u8d28\u6e17\u900f\u7387(m\u00b2)"), ("k10", "\u88c2\u7f1d\u6e17\u900f\u7387(m\u00b2)"),
                ]),
                ("\u591a\u91cd\u5206\u5f62", [
                    ("mf_D0", "D(0)\u5bb9\u91cf\u7ef4\u6570"), ("mf_D1", "D(1)\u4fe1\u606f\u7ef4\u6570"),
                    ("mf_D2", "D(2)\u5173\u8054\u7ef4\u6570"), ("mf_delta_alpha", "\u8c31\u5bbd\u5ea6\u0394\u03b1"),
                    ("mf_delta_f", "\u4e0d\u5bf9\u79f0\u6027\u0394f"),
                ]),
            ]
            with dpg.group(horizontal=True, parent="res_header"):
                for sec_name, items in sections:
                    with dpg.group(parent="res_header"):
                        dpg.add_text(sec_name, color=C_CYAN, parent="res_header")
                        for key, lbl in items:
                            with dpg.group(horizontal=True, parent="res_header"):
                                dpg.add_text(f"{lbl}:", color=[180, 180, 195], parent="res_header")
                                t = dpg.add_text("---", color=C_CYAN, tag=f"r_{key}", parent="res_header")
                                self.res_tags[key] = t

    def _create_plot(self):
        with dpg.plot(tag="plt", height=-1, width=-1, no_title=True, parent="plot_area"):
            dpg.add_plot_legend(location=dpg.mvPlot_Location_NorthWest, tag="plt_legend")
            dpg.add_plot_axis(dpg.mvXAxis, label="X", tag="xax")
            self.yax = dpg.add_plot_axis(dpg.mvYAxis, label="Y", tag="yax")
        dpg.bind_item_theme("plt", "_pt")

    def _recreate_plot(self, tid):
        dpg.delete_item("plt")
        xlabel, ylabel, x_log, x_inv = self._chart_axes(tid)
        with dpg.plot(tag="plt", height=-1, width=-1, no_title=True, parent="plot_area"):
            dpg.add_plot_legend(location=dpg.mvPlot_Location_NorthWest, tag="plt_legend")
            dpg.add_plot_axis(dpg.mvXAxis, label=xlabel, tag="xax", scale=1 if x_log else 0, invert=x_inv)
            self.yax = dpg.add_plot_axis(dpg.mvYAxis, label=ylabel, tag="yax")
        dpg.bind_item_theme("plt", "_pt")
        if self.result:
            self._draw(tid)

    def _chart_axes(self, tid):
        axes = {
            "capillary":      ("\u5b54\u5589\u76f4\u5f84 (nm)", "\u6c55\u9971\u548c\u5ea6 (%)",       True,  True),
            "dvdD":           ("\u5b54\u5589\u76f4\u5f84 (nm)", "dV/dD (mL/g\u00b7nm)", True, True),
            "dvdlogD":        ("\u5b54\u5589\u76f4\u5f84 (nm)", "dV/dlogD (mL/g)",    True,  True),
            "pct":            ("\u5b54\u5589\u76f4\u5f84 (nm)", "\u8fdb\u6c55\u91cf (%)",         True,  True),
            "inj_ext":        ("\u538b\u529b (MPa)",     "\u7d2f\u79ef\u8fdb\u6c55\u91cf (mL/g)",  True,  False),
            "characteristic": ("\u5b54\u5589\u76f4\u5f84 (\u03bcm)", "\u7d2f\u79ef\u9971\u548c\u5ea6 (%)", True,  True),
            "fractal":        ("log(P)",         "log(1-S)",           False, False),
            "mf_spectrum":    ("\u03b1",          "f(\u03b1)",           False, False),
            "mf_Dq":          ("q",              "D(q)",               False, False),
            "mf_tau":         ("q",              "\u03c4(q)",           False, False),
            "ratio":          ("\u6c55\u9971\u548c\u5ea6",       "\u5b54\u5589\u6bd4",             False, False),
            "swanson":        ("\u5b54\u5589\u76f4\u5f84 (\u03bcm)", "S\u00d7D\u00b3", True, True),
        }
        return axes.get(tid, ("X", "Y", False, False))

    def _on_chart_sel(self, sender, app_data):
        for c in CHARTS:
            if c[1] == app_data:
                self.current_chart = c[0]
                break
        self._recreate_plot(self.current_chart)

    def _color_series(self, series_tag, color):
        with dpg.theme() as theme:
            with dpg.theme_component(dpg.mvLineSeries):
                dpg.add_theme_color(dpg.mvPlotCol_Line, [*color, 255])
            with dpg.theme_component(dpg.mvScatterSeries):
                dpg.add_theme_color(dpg.mvPlotCol_MarkerFill, [*color, 255])
                dpg.add_theme_color(dpg.mvPlotCol_MarkerOutline, [*color, 255])
            with dpg.theme_component(dpg.mvBarSeries):
                dpg.add_theme_color(dpg.mvPlotCol_Fill, [*color, 200])
        dpg.bind_item_theme(series_tag, theme)

    def _safe(self, x, y, positive_x=False, nonneg_y=False):
        v = np.isfinite(x) & np.isfinite(y)
        if positive_x:
            v = v & (x > 0)
        if nonneg_y:
            v = v & (y >= 0)
        return x[v].tolist(), y[v].tolist()

    def _update_charts(self):
        self._recreate_plot(self.current_chart)

    def _draw(self, tid):
        r = self.result
        if not r or len(r.cum_intrusion) < 2:
            return
        dispatch = {
            "capillary": self._draw_capillary, "dvdD": self._draw_dvdD,
            "dvdlogD": self._draw_dvdlogD, "pct": self._draw_pct,
            "inj_ext": self._draw_inj_ext, "characteristic": self._draw_characteristic,
            "fractal": self._draw_fractal, "ratio": self._draw_ratio,
            "swanson": self._draw_swanson,
            "mf_spectrum": self._draw_mf_spectrum, "mf_Dq": self._draw_mf_Dq,
            "mf_tau": self._draw_mf_tau,
        }
        fn = dispatch.get(tid)
        if fn:
            fn(r)

    def _draw_capillary(self, r):
        cum_max = np.max(r.cum_intrusion)
        sat = r.cum_intrusion / cum_max * 100 if cum_max > 0 else r.cum_intrusion
        x, y = self._safe(r.pore_throat_diameter_nm, sat, positive_x=True, nonneg_y=True)
        if x:
            s = dpg.add_line_series(x, y, label="\u8fdb\u6c55", parent=self.yax)
            self._color_series(s, C_CYAN)
        if len(r.extrusion_pressure_mpa) > 0 and len(r.cum_extrusion) > 0:
            sigma = self.processor.params.surface_tension
            theta = np.radians(self.processor.params.contact_angle)
            ext_d = -2 * sigma * np.cos(theta) / r.extrusion_pressure_mpa * 2 * 1000
            mask = ext_d > 0
            ext_d_valid = ext_d[mask]
            ext_sat = r.cum_extrusion[mask] / cum_max * 100 if cum_max > 0 else r.cum_extrusion[mask]
            if len(ext_d_valid) > 0:
                x2, y2 = self._safe(ext_d_valid, ext_sat, positive_x=True, nonneg_y=True)
                if x2:
                    s2 = dpg.add_line_series(x2, y2, label="\u9000\u6c55", parent=self.yax)
                    self._color_series(s2, C_RED)

    def _draw_dvdD(self, r):
        x, y = self._safe(r.pore_throat_diameter_nm, r.dv_dD, positive_x=True, nonneg_y=True)
        if x:
            s = dpg.add_line_series(x, y, label="dV/dD", parent=self.yax)
            self._color_series(s, C_GOLD)

    def _draw_dvdlogD(self, r):
        x, y = self._safe(r.pore_throat_diameter_nm, r.dv_dlogD, positive_x=True, nonneg_y=True)
        if x:
            s = dpg.add_line_series(x, y, label="dV/dlogD", parent=self.yax)
            self._color_series(s, C_RED)

    def _draw_pct(self, r):
        if len(r.pore_throat_bins) == 0:
            return
        x, y = self._safe(r.pore_throat_bins, r.bin_pct, positive_x=True, nonneg_y=True)
        if x:
            s = dpg.add_bar_series(x, y, label="\u8fdb\u6c55\u91cf(%)", parent=self.yax, weight=0.8)
            self._color_series(s, C_PURPLE)

    def _draw_inj_ext(self, r):
        x, y = self._safe(r.intrusion_pressure_mpa, r.cum_intrusion)
        if x:
            s = dpg.add_line_series(x, y, label="\u8fdb\u6c55", parent=self.yax)
            self._color_series(s, C_CYAN)
        if len(r.extrusion_pressure_mpa) > 0 and len(r.cum_extrusion) > 0:
            x2, y2 = self._safe(r.extrusion_pressure_mpa, r.cum_extrusion)
            if x2:
                s2 = dpg.add_line_series(x2, y2, label="\u9000\u6c55", parent=self.yax)
                self._color_series(s2, C_RED)

    def _draw_characteristic(self, r):
        cum_max = np.max(r.cum_intrusion)
        if cum_max <= 0:
            return
        sat = r.cum_intrusion / cum_max * 100
        d_um = r.pore_throat_diameter_um
        x1, y1 = self._safe(d_um, sat, positive_x=True, nonneg_y=True)
        if x1:
            s = dpg.add_line_series(x1, y1, label="\u7d2f\u79ef\u9971\u548c\u5ea6", parent=self.yax)
            self._color_series(s, C_CYAN)

    def _draw_fractal(self, r):
        valid = np.isfinite(r.logP) & np.isfinite(r.log1_S) & (r.intrusion_pressure_mpa > 0)
        x = r.logP[valid].tolist()
        y = r.log1_S[valid].tolist()
        if x:
            s = dpg.add_scatter_series(x, y, label="\u6570\u636e\u70b9", parent=self.yax, size=3)
            self._color_series(s, C_CYAN)
        if r.fractal_slopes and len(r.fractal_slopes) > 0 and r.fractal_slopes[0] != 0:
            lp_v = r.logP[valid]
            slope = r.fractal_slopes[0]
            intercept = float(np.mean(r.log1_S[valid] - slope * r.logP[valid]))
            x_fit = np.linspace(float(lp_v.min()), float(lp_v.max()), 100).tolist()
            y_fit = [slope * xi + intercept for xi in x_fit]
            d_label = r.fractal_dimensions[0] if r.fractal_dimensions else 0
            s2 = dpg.add_line_series(x_fit, y_fit, label=f"\u62df\u5408 (D={d_label:.3f})", parent=self.yax)
            self._color_series(s2, C_RED)

    def _draw_ratio(self, r):
        if r.pore_throat_ratio_data:
            d = r.pore_throat_ratio_data
            x, y = self._safe(np.array(d['threshold']), np.array(d['pore_throat_ratio']))
            if x:
                s = dpg.add_line_series(x, y, label="\u5b54\u5589\u6bd4", parent=self.yax)
                self._color_series(s, C_GOLD)

    def _draw_swanson(self, r):
        d_um = r.pore_throat_diameter_nm * 0.001
        cum_max = np.max(r.cum_intrusion)
        sat = r.cum_intrusion / cum_max if cum_max > 0 else r.cum_intrusion
        product = sat * d_um ** 3
        x, y = self._safe(d_um, product, positive_x=True, nonneg_y=True)
        if x:
            s = dpg.add_line_series(x, y, label="S\u00d7D\u00b3", parent=self.yax)
            self._color_series(s, C_PURPLE)
        if len(product) > 0:
            idx = int(np.argmax(product))
            px, py = d_um[idx], product[idx]
            if np.isfinite(px) and np.isfinite(py):
                s2 = dpg.add_scatter_series([px], [py], label=f"\u5cf0\u503c D={px:.3f}\u03bcm", parent=self.yax, size=10)
                self._color_series(s2, C_RED)

    # ---- 多重分形图表 ----
    def _draw_mf_spectrum(self, r):
        if len(r.mf_alpha) < 2:
            return
        valid = np.isfinite(r.mf_alpha) & np.isfinite(r.mf_falpha)
        x, y = self._safe(r.mf_alpha[valid], r.mf_falpha[valid])
        if x:
            s = dpg.add_scatter_series(x, y, label="f(\u03b1)", parent=self.yax, size=5)
            self._color_series(s, C_CYAN)
        # 标注关键点
        if r.mf_D0 > 0:
            idx0 = np.argmin(np.abs(r.mf_q))
            if idx0 < len(r.mf_alpha) and np.isfinite(r.mf_alpha[idx0]):
                s2 = dpg.add_scatter_series([r.mf_alpha[idx0]], [r.mf_falpha[idx0]],
                                            label=f"D(0)={r.mf_D0:.3f}", parent=self.yax, size=8)
                self._color_series(s2, C_GOLD)

    def _draw_mf_Dq(self, r):
        if len(r.mf_q) < 2:
            return
        valid = np.isfinite(r.mf_Dq)
        x, y = self._safe(r.mf_q[valid], r.mf_Dq[valid])
        if x:
            s = dpg.add_scatter_series(x, y, label="D(q)", parent=self.yax, size=5)
            self._color_series(s, C_RED)
        # D(0) 参考线
        if r.mf_D0 > 0:
            q_valid = r.mf_q[valid]
            dq = np.array([float(np.min(q_valid)), float(np.max(q_valid))])
            dy = np.array([r.mf_D0, r.mf_D0])
            s2 = dpg.add_line_series(dq.tolist(), dy.tolist(),
                                      label=f"D(0)={r.mf_D0:.3f}", parent=self.yax)
            self._color_series(s2, C_GOLD)

    def _draw_mf_tau(self, r):
        if len(r.mf_q) < 2:
            return
        valid = np.isfinite(r.mf_tau_q)
        x, y = self._safe(r.mf_q[valid], r.mf_tau_q[valid])
        if x:
            s = dpg.add_scatter_series(x, y, label="\u03c4(q)", parent=self.yax, size=5)
            self._color_series(s, C_PURPLE)

    def _apply_params(self):
        if not self.processor:
            return
        try:
            self.processor.params.contact_angle = dpg.get_value("p_contact_angle")
            self.processor.params.surface_tension = dpg.get_value("p_surface_tension")
            self.processor.params.mercury_injection_pressure = dpg.get_value("p_mercury_pressure")
            self.processor.params.withdrawal_coefficient = dpg.get_value("p_withdrawal_coeff")
            self.processor.params.smoothing_window = int(dpg.get_value("p_smoothing"))
        except Exception:
            pass

    def _apply_params_to(self, proc):
        try:
            proc.params.contact_angle = dpg.get_value("p_contact_angle")
            proc.params.surface_tension = dpg.get_value("p_surface_tension")
            proc.params.mercury_injection_pressure = dpg.get_value("p_mercury_pressure")
            proc.params.withdrawal_coefficient = dpg.get_value("p_withdrawal_coeff")
            proc.params.smoothing_window = int(dpg.get_value("p_smoothing"))
        except Exception:
            pass

    def _update_sample(self):
        d = self.processor.data
        r = self.result
        dpg.set_value("sample_txt",
            f"\u6837\u54c1: {d.sample_name}\n"
            f"\u8d28\u91cf: {d.sample_mass} g\n"
            f"\u63a5\u89e6\u89d2: {self.processor.params.contact_angle}\u00b0\n"
            f"\u6c25\u5b54\u9699\u5ea6: {d.porosity}%\n"
            f"\u8fdb\u6c55: {r.n_intrusion_points}  \u9000\u6c55: {r.n_withdrawal_points}"
        )

    def _update_results(self):
        r = self.result
        if not r:
            return
        vals = {
            'he_por': f"{r._he_porosity:.4f}", 'micp_por': f"{r.cal_porosity:.4f}",
            'bulk_d': f"{r.bulk_density:.4f}", 'skel_d': f"{r.skeletal_density:.4f}",
            'ssa': f"{r.specific_surface_area:.4f}",
            'pore_v': f"{r.pore_volume:.6f}", 'total_pore_area': f"{r.total_pore_area:.4f}",
            'avg_pd': f"{r.avg_pore_diameter_nm:.2f}",
            'inj_sat': f"{r.intrusion_saturation:.2f}", 'eff': f"{r.efficiency:.4f}",
            'disp_p': f"{r.displacement_pressure:.4f}", 'max_d': f"{r.max_pore_diameter_um:.4f}",
            'med_p': f"{r.median_pressure:.4f}", 'med_d': f"{r.median_diameter_um:.6f}",
            'med_dvol': f"{r.median_pore_diameter_volume_nm:.2f}", 'med_darea': f"{r.median_pore_diameter_area_nm:.2f}",
            'sp': f"{r.sorting_coefficient:.4f}", 'skp': f"{r.skewness:.4f}",
            'kp': f"{r.kurtosis:.4f}", 'dm': f"{r.mean_radius:.4f}",
            'phi': f"{r.structure_coefficient:.4f}", 'd_coeff': f"{r.relative_sorting_coeff:.4f}",
            'frac_d': f"{r.fractal_dimensions[0]:.4f}" if r.fractal_dimensions else "---",
            'k413': f"{r.permeability_413:.6e}", 'k10': f"{r.permeability_10:.6e}",
            'bpr': f"{r.breakthrough_pressure_ratio:.4f}", 'cl': f"{r.characteristic_length_nm:.2f}",
            'cff': f"{r.conductivity_formation_factor:.4f}",
            'tf': f"{r.tortuosity_factor:.4f}", 'tort': f"{r.tortuosity:.4f}",
            'mf_D0': f"{r.mf_D0:.4f}", 'mf_D1': f"{r.mf_D1:.4f}",
            'mf_D2': f"{r.mf_D2:.4f}", 'mf_delta_alpha': f"{r.mf_delta_alpha:.4f}",
            'mf_delta_f': f"{r.mf_delta_f:.4f}",
        }
        for k, v in vals.items():
            t = self.res_tags.get(k)
            if t:
                dpg.set_value(t, v)

    def _show_open(self):
        dpg.show_item("fd_open")

    def _show_batch(self):
        dpg.show_item("fd_batch")

    def _on_file(self, sender, app_data):
        fp = app_data['file_path_name']
        try:
            self.processor = MICPProcessor()
            self.processor.load(fp)
            self._apply_params()
            self.result = self.processor.process()
            dpg.set_value("file_lbl", os.path.basename(fp))
            self._update_sample()
            self._update_results()
            self._update_charts()
        except Exception as e:
            import traceback
            traceback.print_exc()
            tag = dpg.generate_uuid()
            with dpg.window(label="\u9519\u8bef", modal=True, tag=tag, no_resize=True, width=420, height=160):
                dpg.add_text(str(e), color=C_RED, wrap=400)
                dpg.add_button(label="\u5173\u95ed", callback=lambda: dpg.delete_item(tag))

    def _on_batch(self, sender, app_data):
        folder = app_data['file_path_name']
        exts = ('.xlsx', '.XLS', '.xls', '.XLSX')
        try:
            files = [f for f in os.listdir(folder)
                     if f.endswith(exts) and not f.startswith('~$')
                     and "template" not in f.lower() and not f.endswith('.xlsm')]
        except Exception as e:
            self._error(str(e))
            return
        if not files:
            return
        self.batch_results = []
        ok = 0
        for fname in files:
            try:
                p = MICPProcessor()
                p.load(os.path.join(folder, fname))
                self._apply_params_to(p)
                r = p.process()
                self.batch_results.append({'name': fname, 'result': r, 'data': p.data, 'processor': p})
                ok += 1
            except Exception as e:
                self.batch_results.append({'name': fname, 'error': str(e)})
        dpg.set_value("file_lbl", f"\u6279\u91cf: {ok}/{len(files)}")
        if ok > 0:
            self.processor = self.batch_results[0]['processor']
            self.result = self.batch_results[0]['result']
            self._update_sample()
            self._update_results()
            self._update_charts()

    def _recalc(self):
        if not self.processor:
            return
        try:
            self._apply_params()
            self.result = self.processor.process()
            self._update_results()
            self._update_charts()
        except Exception as e:
            self._error(str(e))

    def _error(self, msg):
        tag = dpg.generate_uuid()
        with dpg.window(label="\u9519\u8bef", modal=True, tag=tag, no_resize=True, width=420, height=160):
            dpg.add_text(str(msg), color=C_RED, wrap=400)
            dpg.add_button(label="\u5173\u95ed", callback=lambda: dpg.delete_item(tag))

    def _do_export_xlsx(self):
        if not self.result:
            return
        dpg.show_item("fd_sxlsx")

    def _do_export_json(self):
        if not self.result:
            return
        dpg.show_item("fd_sjson")

    def _on_save_xlsx(self, sender, app_data):
        fp = app_data['file_path_name']
        try:
            if len(self.batch_results) > 1:
                self._batch_export_xlsx(fp)
            else:
                export_excel(fp, self.processor.data, self.result)
        except Exception as e:
            self._error(str(e))

    def _on_save_json(self, sender, app_data):
        fp = app_data['file_path_name']
        try:
            export_json(fp, self.processor.data, self.result)
        except Exception as e:
            self._error(str(e))

    def _batch_export_xlsx(self, fp):
        import openpyxl
        from openpyxl.styles import Font, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "\u6c47\u603b"
        border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        headers = [
            "\u6837\u54c1\u540d\u79f0", "\u6c25\u5b54\u9699\u5ea6(%)", "\u538b\u6c55\u5b54\u9699\u5ea6(%)",
            "\u4f53\u79ef\u5bc6\u5ea6", "\u9aa8\u67b6\u5bc6\u5ea6", "\u6bd4\u8868\u9762\u79ef",
            "\u6392\u9a71\u538b\u529b", "\u6700\u5927\u5b54\u5f84(\u03bcm)", "\u4e2d\u503c\u538b\u529b",
            "\u4e2d\u503c\u5b54\u5f84(\u03bcm)", "\u5206\u9009\u7cfb\u6570Sp", "\u6b6a\u5ea6Skp",
            "\u5cf0\u6001Kp", "\u5206\u5f62\u7ef4\u6570", "\u5b54\u9699\u4f53\u79ef"
        ]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True)
            c.border = border
        for ri, item in enumerate(self.batch_results, 2):
            if 'error' in item:
                ws.cell(row=ri, column=1, value=item['name'])
                ws.cell(row=ri, column=2, value="ERROR")
                continue
            rep = item['result'].to_report_dict()
            for ci, k in enumerate([
                'sample_name', 'he_porosity', 'micp_porosity',
                'bulk_density', 'skeletal_density', 'specific_surface_area',
                'displacement_pressure', 'max_pore_diameter_um',
                'median_pressure', 'median_diameter_um',
                'sorting_coefficient', 'skewness', 'kurtosis',
                'fractal_dimension', 'pore_volume'
            ], 1):
                ws.cell(row=ri, column=ci, value=rep.get(k, '')).border = border
        wb.save(fp)


def main():
    app = App()
    app.run()


if __name__ == "__main__":
    main()
