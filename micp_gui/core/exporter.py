import json
import numpy as np
from datetime import datetime

from .models import MICPData, MICPResult


def export_excel(file_path: str, data: MICPData, result: MICPResult):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    hfont = Font(bold=True, size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont_w = Font(bold=True, size=11, color="FFFFFF")
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal='center')

    report = result.to_report_dict()

    ws.merge_cells('A1:D1')
    ws['A1'] = f"压汞法毛管压力曲线测定报告 - {report['sample_name']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    row = 3
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="基本信息").font = Font(bold=True, size=12)
    row += 1

    for items in [
        ("样品名称", report['sample_name'], "测试时间", report['test_time']),
        ("氦孔隙度(%)", f"{report['he_porosity']:.4f}", "压汞孔隙度(%)", f"{report['micp_porosity']:.4f}"),
        ("体积密度(g/cm³)", f"{report['bulk_density']:.4f}", "骨架密度(g/cm³)", f"{report['skeletal_density']:.4f}"),
        ("比表面积(m²/g)", f"{report['specific_surface_area']:.4f}", "", ""),
    ]:
        for ci, val in enumerate(items, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = border
            if ci in [1, 3]:
                c.font = hfont
        row += 1

    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="结构参数").font = Font(bold=True, size=12)
    row += 1

    for items in [
        ("进汞饱和度(%)", f"{report['intrusion_saturation']:.2f}", "退汞效率(%)", f"{report['efficiency']:.4f}"),
        ("排驱压力(MPa)", f"{report['displacement_pressure']:.4f}", "最大孔径(μm)", f"{report['max_pore_diameter_um']:.4f}"),
        ("中值压力(MPa)", f"{report['median_pressure']:.4f}", "中值孔径(μm)", f"{report['median_diameter_um']:.6f}"),
        ("中值孔径体积(nm)", f"{report.get('median_pore_diameter_volume_nm', 0):.2f}", "中值孔径面积(nm)", f"{report.get('median_pore_diameter_area_nm', 0):.2f}"),
        ("平均孔径(nm)", f"{report.get('avg_pore_diameter_nm', 0):.2f}", "总孔面积(m²/g)", f"{report.get('total_pore_area', 0):.4f}"),
        ("基质渗透率(m²)", f"{report['matrix_permeability']:.6e}", "裂缝渗透率(m²)", f"{report['fracture_permeability']:.6e}"),
        ("孔隙体积(mL/g)", f"{report['pore_volume']:.6f}", "", ""),
    ]:
        for ci, val in enumerate(items, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = border
            if ci in [1, 3]:
                c.font = hfont
        row += 1

    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="特征参数").font = Font(bold=True, size=12)
    row += 1

    for items in [
        ("分选系数Sp", f"{report['sorting_coefficient']:.4f}", "歪度Skp", f"{report['skewness']:.4f}"),
        ("峰态Kp", f"{report['kurtosis']:.4f}", "半径均值DM", f"{report['mean_radius']:.4f}"),
        ("结构系数ϕ", f"{report['structure_coefficient']:.4f}", "相对分选系数D", f"{report['relative_sorting_coeff']:.4f}"),
        ("分形维数", f"{report['fractal_dimension']:.4f}", "逾渗分形维数", f"{report.get('percolation_fractal_dimension', 0):.4f}"),
        ("特征长度(nm)", f"{report.get('characteristic_length_nm', 0):.2f}", "导电地层因子", f"{report.get('conductivity_formation_factor', 0):.4f}"),
        ("迂曲度因子", f"{report.get('tortuosity_factor', 0):.4f}", "迂曲度", f"{report.get('tortuosity', 0):.4f}"),
        ("突破压力比", f"{report.get('breakthrough_pressure_ratio', 0):.4f}", "", ""),
    ]:
        for ci, val in enumerate(items, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = border
            if ci in [1, 3]:
                c.font = hfont
        row += 1

    # ---- 多重分形参数 ----
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="多重分形参数").font = Font(bold=True, size=12)
    row += 1

    for items in [
        ("D(0)", f"{report.get('mf_D0', 0):.4f}", "D(1)", f"{report.get('mf_D1', 0):.4f}"),
        ("D(2)", f"{report.get('mf_D2', 0):.4f}", "Δα", f"{report.get('mf_delta_alpha', 0):.4f}"),
        ("Δf", f"{report.get('mf_delta_f', 0):.4f}", "D_a", f"{report.get('mf_D_a', 0):.4f}"),
        ("R_d", f"{report.get('mf_R_d', 0):.4f}", "D_Fa", f"{report.get('mf_D_Fa', 0):.4f}"),
        ("D(-10)", f"{report.get('mf_D_neg10', 0):.4f}", "D(10)", f"{report.get('mf_D_10', 0):.4f}"),
        ("H", f"{report.get('mf_H', 0):.4f}", "D(-10)-D(10)", f"{report.get('mf_D_neg10_minus_D_10', 0):.4f}"),
    ]:
        for ci, val in enumerate(items, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = border
            if ci in [1, 3]:
                c.font = hfont
        row += 1

    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 22

    ws2 = wb.create_sheet("原始数据")
    for ci, h in enumerate(["压力(MPa)", "累积进汞量(mL/g)", "孔喉直径(nm)", "进汞增量(mL/g)", "dV/dD", "dV/dlogD"], 1):
        ws2.cell(row=1, column=ci, value=h).font = hfont

    for i in range(len(result.intrusion_pressure_mpa)):
        r = i + 2
        ws2.cell(row=r, column=1, value=result.intrusion_pressure_mpa[i])
        ws2.cell(row=r, column=2, value=result.cum_intrusion[i])
        if i < len(result.pore_throat_diameter_nm):
            ws2.cell(row=r, column=3, value=result.pore_throat_diameter_nm[i])
        if i < len(result.incremental_intrusion):
            ws2.cell(row=r, column=4, value=result.incremental_intrusion[i])
        if i < len(result.dv_dD):
            ws2.cell(row=r, column=5, value=result.dv_dD[i])
        if i < len(result.dv_dlogD):
            ws2.cell(row=r, column=6, value=result.dv_dlogD[i])

    wb.save(file_path)


def export_json(file_path: str, data: MICPData, result: MICPResult):
    report = result.to_report_dict()
    export_data = {
        'sample_info': report,
        'raw_data': {
            'pressure_mpa': result.intrusion_pressure_mpa.tolist(),
            'cum_intrusion': result.cum_intrusion.tolist(),
            'pore_throat_diameter_nm': result.pore_throat_diameter_nm.tolist(),
            'incremental_intrusion': result.incremental_intrusion.tolist(),
            'dv_dD': result.dv_dD.tolist(),
            'dv_dlogD': result.dv_dlogD.tolist(),
        },
        'export_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    if len(result.extrusion_pressure_mpa) > 0:
        export_data['raw_data']['extrusion_pressure_mpa'] = result.extrusion_pressure_mpa.tolist()
        export_data['raw_data']['cum_extrusion'] = result.cum_extrusion.tolist()

    if len(result.pore_throat_bins) > 0:
        export_data['binned_psd'] = {
            'pore_diameter_nm': result.pore_throat_bins.tolist(),
            'intrusion_pct': result.bin_pct.tolist(),
        }

    if len(result.mf_q) > 0:
        export_data['multifractal'] = {
            'q': result.mf_q.tolist(),
            'alpha': result.mf_alpha.tolist(),
            'falpha': result.mf_falpha.tolist(),
            'Dq': result.mf_Dq.tolist(),
            'tau_q': result.mf_tau_q.tolist(),
            'D0': result.mf_D0,
            'D1': result.mf_D1,
            'D2': result.mf_D2,
            'D_neg10': result.mf_D_neg10,
            'D_10': result.mf_D_10,
            'D_neg10_minus_D_10': result.mf_D_neg10_minus_D_10,
            'H': result.mf_H,
            'delta_alpha': result.mf_delta_alpha,
            'delta_f': result.mf_delta_f,
            'D_a': result.mf_D_a,
            'R_d': result.mf_R_d,
            'D_Fa': result.mf_D_Fa,
            'a_min': result.mf_a_min,
            'a_max': result.mf_a_max,
            'F_max': result.mf_F_max,
            'F_min': result.mf_F_min,
        }

    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)
